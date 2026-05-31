"""
FDLE property crime — Lee County + Collier County ingest pipeline.

Two data sources:
  FIBRS (2021–present):  Florida Incident-Based Reporting System, one Excel file
    with one sheet per year, County+Agency rows, 12 monthly columns per offense type.
    Offense groups: Burglary, all Larceny-* subtypes, Motor Vehicle Theft, Arson.
    URL: FIBRS_URL in constants.py.  Verify at FIBRS_CITATION_URL if 404.

  UCR county property crime (2010–2020): FDLE static content-asset Excel files,
    one per year, county-level annual totals (one row per county).
    URLs: UCR_COUNTY_URLS dict in constants.py.

Tiers:
  Tier 1 — Raw rows → NDJSON → Supabase Storage
             path: lake-tier1/crime/{year}/fdle_crime_swfl.ndjson
  Tier 2 — Aggregated rows → PostgreSQL public.fdle_crime_swfl
             upsert key: (county, period)

Cadence: quarterly GHA cron picks up each new annual FDLE release.

Usage:
  python -m ingest.pipelines.fdle_crime_swfl.pipeline [--backfill] [--current]
      [--year YYYY] [--dry-run]

Environment:
  DESTINATION__POSTGRES__CREDENTIALS  — psycopg3 connection URI
  SUPABASE_URL                        — Supabase project URL (Tier-1 upload)
  SUPABASE_SERVICE_KEY                — Supabase service role key (Tier-1 upload)
"""
from __future__ import annotations

import argparse
import io
import json
import os
import sys
import time
from datetime import datetime, timezone
from typing import Any

import openpyxl
import psycopg
import requests

from ingest.lib.tier1_inventory import upsert_inventory_row

from .constants import (
    COL_ALIASES,
    COUNTIES,
    EARLIEST_YEAR,
    FIBRS_CITATION_URL,
    FIBRS_FIRST_YEAR,
    FIBRS_OFFENSE_GROUPS,
    FIBRS_URL,
    TABLE,
    TIER1_BUCKET,
    TIER1_PREFIX,
    UCR_CITATION_URL,
    UCR_COUNTY_URLS,
)

Row = dict[str, Any]


# ── Helpers ───────────────────────────────────────────────────────────────────


def _current_year() -> int:
    """Most recent year for which FDLE typically has data (prior calendar year)."""
    return datetime.now(timezone.utc).year - 1


def _to_int(value: Any) -> int | None:
    if value is None:
        return None
    try:
        return int(float(str(value).replace(",", "").strip()))
    except (ValueError, TypeError):
        return None


def _normalize_cell(s: object) -> str:
    return str(s or "").strip().lower().replace("  ", " ")


def _download(url: str, label: str, timeout: int = 120) -> bytes:
    for attempt in range(3):
        try:
            resp = requests.get(url, timeout=timeout)
            if resp.status_code == 404:
                raise FileNotFoundError(f"{label} not found at {url}")
            resp.raise_for_status()
            return resp.content
        except (requests.Timeout, requests.ConnectionError):
            if attempt < 2:
                time.sleep(5 * (attempt + 1))
                continue
            raise
    raise RuntimeError(f"Failed to download {label} after 3 attempts")


# ── Tier-1 NDJSON upload ──────────────────────────────────────────────────────


def _upload_ndjson(bucket: str, object_path: str, rows: list[Row]) -> int:
    ndjson_bytes = "\n".join(json.dumps(r, default=str) for r in rows).encode("utf-8")
    url = f"{os.environ['SUPABASE_URL']}/storage/v1/object/{bucket}/{object_path}"
    headers = {
        "Authorization": f"Bearer {os.environ['SUPABASE_SERVICE_KEY']}",
        "Content-Type": "application/x-ndjson",
        "x-upsert": "true",
    }
    for attempt in range(3):
        try:
            resp = requests.post(url, headers=headers, data=ndjson_bytes, timeout=60)
            if resp.ok:
                return len(ndjson_bytes)
            if resp.status_code >= 500 and attempt < 2:
                time.sleep(10 * (attempt + 1))
                continue
            raise RuntimeError(
                f"Storage upload failed {resp.status_code}: {resp.text[:200]}"
            )
        except (requests.Timeout, requests.ConnectionError):
            if attempt < 2:
                time.sleep(5 * (attempt + 1))
                continue
            raise
    return len(ndjson_bytes)


# ── FIBRS parser (2021–present) ───────────────────────────────────────────────


def _parse_fibrs_sheet(ws: Any, year: int) -> list[Row]:
    """Parse one FIBRS year sheet into county-aggregated annual rows.

    FIBRS layout (3 header rows):
      Row 1: year title string
      Row 2: offense group headers at the start column of each 12-month block
             (e.g., "Burglary", "Larceny - Pocket Picking", "Motor Vehicle Theft", "Arson")
      Row 3: data column labels — "County", "Agency Name", then monthly "Jan"/"Feb"/...
             repeated for each offense group

    Column layout:
      Col 0: County
      Col 1: Agency Name
      Col 2: Population Estimate (single value, not repeated)
      Col 3+: offense groups, each 12 columns wide (Jan–Dec)

    We aggregate all agency rows per county and sum annual totals per offense.
    All "Larceny - *" subtypes are summed into larceny_theft.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < 4:
        return []

    # Row 2 (index 1): offense group headers — find start columns for each slug
    offense_row = all_rows[1]
    # Map slug → list of (start_col_index,) for that group's 12-month block
    group_cols: dict[str, list[int]] = {slug: [] for slug in FIBRS_OFFENSE_GROUPS}

    for col_idx, cell in enumerate(offense_row):
        if cell is None:
            continue
        cell_norm = str(cell).strip().lower()
        for slug, keywords in FIBRS_OFFENSE_GROUPS.items():
            if any(kw in cell_norm for kw in keywords):
                group_cols[slug].append(col_idx)

    # Population is at col 2 (row 2 label: " {year} Population Estimate")
    POP_COL = 2
    COUNTY_COL = 0

    # Accumulate: county → slug → total count
    county_totals: dict[str, dict[str, int]] = {c: {s: 0 for s in FIBRS_OFFENSE_GROUPS} for c in COUNTIES}
    county_population: dict[str, int | None] = {c: None for c in COUNTIES}

    # Data rows start at index 3 (after 3 header rows)
    for data_row in all_rows[3:]:
        if not data_row or data_row[COUNTY_COL] is None:
            continue
        county_raw = _normalize_cell(data_row[COUNTY_COL])
        matched: str | None = None
        for target in COUNTIES:
            if county_raw.startswith(target.lower()):
                matched = target
                break
        if matched is None:
            continue

        # Population: take the first non-None value seen for this county
        if county_population[matched] is None and len(data_row) > POP_COL:
            pop_val = _to_int(data_row[POP_COL])
            if pop_val:
                county_population[matched] = pop_val

        # Sum 12 monthly columns for each offense group
        for slug, start_cols in group_cols.items():
            for start_col in start_cols:
                for month_offset in range(12):
                    col = start_col + month_offset
                    if col < len(data_row):
                        v = _to_int(data_row[col])
                        if v is not None:
                            county_totals[matched][slug] += v

    now_iso = datetime.now(timezone.utc).isoformat()
    rows: list[Row] = []
    for county in COUNTIES:
        totals = county_totals[county]
        # Only emit if we found at least one offense value
        if not any(totals.values()):
            continue
        burglary = totals["burglary"] or None
        larceny_theft = totals["larceny_theft"] or None
        motor_vehicle_theft = totals["motor_vehicle_theft"] or None
        arson = totals["arson"] or None
        parts = [x for x in [burglary, larceny_theft, motor_vehicle_theft, arson] if x is not None]
        total_property_crimes = sum(parts) if parts else None
        population = county_population[county]
        property_crime_per_1k: float | None = None
        if population and total_property_crimes:
            property_crime_per_1k = round(total_property_crimes / population * 1000, 2)
        rows.append({
            "county": county,
            "period": f"{year}-01-01",
            "data_year": year,
            "burglary": burglary,
            "larceny_theft": larceny_theft,
            "motor_vehicle_theft": motor_vehicle_theft,
            "arson": arson,
            "total_property_crimes": total_property_crimes,
            "population": population,
            "property_crime_per_1k": property_crime_per_1k,
            "source_url": FIBRS_URL,
            "retrieved_at": now_iso,
        })
    return rows


def parse_fibrs(content: bytes, years: list[int]) -> dict[int, list[Row]]:
    """Parse all requested years from the FIBRS Excel file.

    Returns {year: [rows]} for years found in the file.
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    results: dict[int, list[Row]] = {}
    for sheet_name in wb.sheetnames:
        # Sheet names: "FIBRS Offense 2021", "FIBRS Offense 2022", etc.
        parts = sheet_name.strip().rsplit(" ", 1)
        if len(parts) != 2:
            continue
        try:
            sheet_year = int(parts[1])
        except ValueError:
            continue
        if sheet_year not in years:
            continue
        ws = wb[sheet_name]
        rows = _parse_fibrs_sheet(ws, sheet_year)
        if rows:
            results[sheet_year] = rows
    wb.close()
    return results


# ── UCR county property crime parser (2010–2020) ──────────────────────────────


def _find_ucr_header_row(ws: Any) -> tuple[int, dict[str, int]]:
    """Scan rows for the UCR county file header (has 'county' + 'population').

    Returns (1-based row index, {slug: 0-based column index}).
    """
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        cells = [_normalize_cell(c) for c in row]
        if not (any("county" in c for c in cells) and
                any("population" in c or "pop." in c for c in cells)):
            continue
        col_map: dict[str, int] = {}
        for slug, aliases in COL_ALIASES.items():
            for col_idx, cell in enumerate(cells):
                if any(alias in cell for alias in aliases):
                    if slug not in col_map:
                        col_map[slug] = col_idx
        if "county" in col_map and "population" in col_map:
            return row_idx, col_map
    raise ValueError(
        "No header row with 'county' and 'population' found. "
        f"Check {UCR_CITATION_URL} — FDLE may have changed the Excel layout."
    )


def parse_ucr_excel(content: bytes, year: int, source_url: str) -> list[Row]:
    """Parse a UCR county property crime Excel file for Lee + Collier."""
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    now_iso = datetime.now(timezone.utc).isoformat()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        try:
            header_row_idx, col_map = _find_ucr_header_row(ws)
        except ValueError:
            continue

        rows: list[Row] = []
        for data_row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            county_key = col_map.get("county")
            if county_key is None:
                continue
            county_raw = _normalize_cell(data_row[county_key])
            if not county_raw:
                continue

            matched_county: str | None = None
            for target in COUNTIES:
                if county_raw.startswith(target.lower()):
                    matched_county = target
                    break
            if matched_county is None:
                continue

            def get_int(slug: str) -> int | None:
                idx = col_map.get(slug)
                if idx is None:
                    return None
                return _to_int(data_row[idx])

            population = get_int("population")
            burglary = get_int("burglary")
            larceny_theft = get_int("larceny_theft")
            motor_vehicle_theft = get_int("motor_vehicle_theft")
            arson = get_int("arson")

            total_from_excel = get_int("total_property")
            parts = [x for x in [burglary, larceny_theft, motor_vehicle_theft, arson] if x is not None]
            total_property_crimes = total_from_excel or (sum(parts) if parts else None)

            property_crime_per_1k: float | None = None
            if population and total_property_crimes:
                property_crime_per_1k = round(total_property_crimes / population * 1000, 2)

            rows.append({
                "county": matched_county,
                "period": f"{year}-01-01",
                "data_year": year,
                "burglary": burglary,
                "larceny_theft": larceny_theft,
                "motor_vehicle_theft": motor_vehicle_theft,
                "arson": arson,
                "total_property_crimes": total_property_crimes,
                "population": population,
                "property_crime_per_1k": property_crime_per_1k,
                "source_url": source_url,
                "retrieved_at": now_iso,
            })

        if rows:
            wb.close()
            return rows

    wb.close()
    raise ValueError(
        f"UCR {year}: no rows found for Lee or Collier. "
        "Verify Excel structure — check UCR_COUNTY_URLS in constants.py."
    )


# ── DB upsert ─────────────────────────────────────────────────────────────────

UPSERT_SQL = f"""
INSERT INTO {TABLE}
  (county, period, data_year, burglary, larceny_theft, motor_vehicle_theft, arson,
   total_property_crimes, population, property_crime_per_1k, source_url, retrieved_at, inserted_at)
VALUES
  (%(county)s, %(period)s::date, %(data_year)s, %(burglary)s, %(larceny_theft)s,
   %(motor_vehicle_theft)s, %(arson)s, %(total_property_crimes)s, %(population)s,
   %(property_crime_per_1k)s, %(source_url)s, %(retrieved_at)s, NOW())
ON CONFLICT (county, period) DO UPDATE SET
  data_year             = EXCLUDED.data_year,
  burglary              = EXCLUDED.burglary,
  larceny_theft         = EXCLUDED.larceny_theft,
  motor_vehicle_theft   = EXCLUDED.motor_vehicle_theft,
  arson                 = EXCLUDED.arson,
  total_property_crimes = EXCLUDED.total_property_crimes,
  population            = EXCLUDED.population,
  property_crime_per_1k = EXCLUDED.property_crime_per_1k,
  source_url            = EXCLUDED.source_url,
  retrieved_at          = EXCLUDED.retrieved_at,
  inserted_at           = NOW()
"""


def upsert_rows(rows: list[Row], conn_str: str) -> int:
    if not rows:
        return 0
    with psycopg.connect(conn_str) as conn:
        with conn.cursor() as cur:
            cur.executemany(UPSERT_SQL, rows)
        conn.commit()
    return len(rows)


# ── Orchestration ─────────────────────────────────────────────────────────────


def run(years: list[int], dry_run: bool, conn_str: str | None) -> None:
    fibrs_years = sorted(y for y in years if y >= FIBRS_FIRST_YEAR)
    ucr_years = sorted(y for y in years if y < FIBRS_FIRST_YEAR and y in UCR_COUNTY_URLS)
    skipped = [y for y in years if y < FIBRS_FIRST_YEAR and y not in UCR_COUNTY_URLS]
    if skipped:
        print(f"  SKIP years with no UCR URL: {skipped}", file=sys.stderr)

    all_year_rows: dict[int, list[Row]] = {}

    # ── FIBRS batch download (one file covers all 2021+ years) ──────────────
    if fibrs_years:
        print(f"fdle_crime_swfl: downloading FIBRS file for years {fibrs_years}...")
        try:
            fibrs_content = _download(FIBRS_URL, "FIBRS offense data")
            fibrs_results = parse_fibrs(fibrs_content, fibrs_years)
            for year, rows in fibrs_results.items():
                all_year_rows[year] = rows
                print(f"  FIBRS {year}: {len(rows)} county rows parsed ({[r['county'] for r in rows]}).")
            missing = [y for y in fibrs_years if y not in fibrs_results]
            if missing:
                print(f"  FIBRS: no sheet found for years {missing} — FDLE may not have published yet.", file=sys.stderr)
        except FileNotFoundError as exc:
            print(f"  SKIP FIBRS: {exc}", file=sys.stderr)
        except Exception as exc:
            print(f"  SKIP FIBRS: {exc}", file=sys.stderr)

    # ── UCR per-year downloads (2010–2020) ───────────────────────────────────
    for year in ucr_years:
        url = UCR_COUNTY_URLS[year]
        print(f"fdle_crime_swfl: downloading UCR {year}...")
        try:
            content = _download(url, f"UCR {year}")
            rows = parse_ucr_excel(content, year, source_url=url)
            all_year_rows[year] = rows
            print(f"  UCR {year}: {len(rows)} county rows parsed ({[r['county'] for r in rows]}).")
        except FileNotFoundError as exc:
            print(f"  SKIP {year}: {exc}", file=sys.stderr)
        except Exception as exc:
            print(f"  SKIP {year}: {exc}", file=sys.stderr)

    # ── Emit / write ─────────────────────────────────────────────────────────
    total = 0
    for year in sorted(all_year_rows):
        rows = all_year_rows[year]
        source_url = rows[0]["source_url"] if rows else FIBRS_URL
        citation_url = FIBRS_CITATION_URL if year >= FIBRS_FIRST_YEAR else UCR_CITATION_URL

        if dry_run:
            for r in rows:
                print(f"  {r}")
            continue

        if not conn_str:
            raise RuntimeError("DESTINATION__POSTGRES__CREDENTIALS not set.")

        supabase_url = os.environ.get("SUPABASE_URL")
        supabase_key = os.environ.get("SUPABASE_SERVICE_KEY")
        if supabase_url and supabase_key:
            ndjson_path = f"{TIER1_PREFIX}/{year}/fdle_crime_swfl.ndjson"
            byte_size = _upload_ndjson(TIER1_BUCKET, ndjson_path, rows)
            upsert_inventory_row(
                bucket=TIER1_BUCKET,
                path=ndjson_path,
                vintage=str(year),
                byte_size=byte_size,
                pack_id="safety-swfl",
                source_url=citation_url,
            )
            print(f"  Tier 1: {byte_size} bytes → {TIER1_BUCKET}/{ndjson_path}")
        else:
            print("  WARNING: SUPABASE_URL/KEY not set — skipping Tier-1 upload.", file=sys.stderr)

        written = upsert_rows(rows, conn_str)
        print(f"  Tier 2: upserted {written} rows.")
        total += written

    if not dry_run:
        print(f"Done. Total rows upserted: {total}.")


# ── CLI ───────────────────────────────────────────────────────────────────────


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="FDLE property crime ingest — Lee + Collier counties."
    )
    mode = parser.add_mutually_exclusive_group()
    mode.add_argument(
        "--backfill",
        action="store_true",
        help=f"Ingest all years from {EARLIEST_YEAR} through current.",
    )
    mode.add_argument(
        "--current",
        action="store_true",
        help="Ingest current and prior year (default for cron).",
    )
    mode.add_argument(
        "--year",
        type=int,
        metavar="YYYY",
        help="Ingest a specific year, e.g. --year 2023.",
    )
    parser.add_argument("--dry-run", action="store_true")

    args = parser.parse_args(argv)
    conn_str = os.environ.get("DESTINATION__POSTGRES__CREDENTIALS")
    current = _current_year()

    if args.backfill:
        years = list(range(EARLIEST_YEAR, current + 1))
    elif args.year:
        years = [args.year]
    else:
        years = [current - 1, current]

    print(
        f"fdle_crime_swfl: {'dry-run ' if args.dry_run else ''}"
        f"years={years}, counties={COUNTIES}"
    )
    run(years, dry_run=args.dry_run, conn_str=conn_str)
    return 0


if __name__ == "__main__":
    sys.exit(main())
