"""
FL DOR Tourist Development Tax — Python ingest pipeline.

Downloads Form 3 Excel from Florida DOR (one file per FL fiscal year), parses
the "Tourist Development Tax" sheet, and upserts Lee + Collier monthly
collections into public.fl_dor_tdt_collections.

FL fiscal year convention:
  FY YYYY covers Jul (YYYY-1) through Jun YYYY.
  FY2026 = Jul 2025 → Jun 2026.
  URL: https://floridarevenue.com/dataPortal/GTA/Form%203/F3FY{YYYY}.xlsx
  Confirmed live FY1999–FY2026.

Excel structure (sheet "Tourist Development Tax"):
  Row 9   — date headers: col A blank, cols B–M = Jul→Jun datetimes.
  Rows 12+— one row per county; col A = "NN*CountyName" pattern.
            Lee = "46*Lee" (~row 47); Collier = "21*Collier" (~row 22).
  Cols B–M— individual monthly USD (not cumulative). None = not yet published.

Lee County note: The Premise data inventory noted "Lee self-administers — FL DOR
has zero Lee data." The plan authors observed Lee at row 47 of Form 3. If Lee is
absent or all-zero the pipeline logs a clear warning and writes Collier only.

Usage:
  python -m ingest.pipelines.fl_dor_tdt.pipeline [--backfill] [--current] [--fy YYYY]
                                                   [--dry-run] [--counties Lee,Collier]

Environment:
  DESTINATION__POSTGRES__CREDENTIALS — psycopg3 connection URI
    (e.g. postgresql://postgres:[pass]@db.[ref].supabase.co:5432/postgres)
"""

from __future__ import annotations

import argparse
import io
import os
import sys
from datetime import datetime, date, timezone
from typing import Iterator

import psycopg
import requests
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

# ── Constants ────────────────────────────────────────────────────────────────

FORM3_URL = "https://floridarevenue.com/dataPortal/GTA/Form%203/F3FY{fy}.xlsx"
SHEET_NAME = "Tourist Development Tax"
TABLE = "fl_dor_tdt_collections"

# Counties we care about and their FIPS codes.
DEFAULT_COUNTIES = ["Lee", "Collier"]
COUNTY_FIPS: dict[str, str] = {
    "Lee": "12071",
    "Collier": "12021",
}

# Source URL template (human-readable, stored per row for citation).
SOURCE_URL_TMPL = "https://floridarevenue.com/dataPortal/GTA/Form%203/F3FY{fy}.xlsx"

# ── Fiscal year helpers ───────────────────────────────────────────────────────


def current_fl_fy(now: datetime | None = None) -> int:
    """Current FL fiscal year. FY YYYY starts Jul 1 of (YYYY-1)."""
    d = now or datetime.now(timezone.utc)
    return d.year + 1 if d.month >= 7 else d.year


def fy_range(start_fy: int, end_fy: int) -> list[int]:
    return list(range(start_fy, end_fy + 1))


# ── Excel parsing ─────────────────────────────────────────────────────────────


def _find_county_row(ws: Worksheet, county: str) -> int | None:
    """Return 1-based row index for the given county, or None if not found."""
    for row_idx in range(12, (ws.max_row or 100) + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val is None:
            continue
        text = str(cell_val).strip()
        # Match "NN*CountyName" or bare county name (case-insensitive).
        if f"*{county.lower()}" in text.lower() or text.lower() == county.lower():
            return row_idx
    return None


def _fy_month_date(fy: int, col_idx: int) -> date:
    """Return the first-of-month date for a Form 3 column.

    FL pipeline FY convention: FY YYYY = Jul (YYYY-1) → Jun YYYY.
    Column indices 0-11 map to Jul, Aug, ..., Dec of (fy-1), then Jan, ..., Jun of fy.
    This bypasses row-9 cell parsing — old Excel files (FY1999, 2001-2003) store
    month headers as small serials that openpyxl reads as 1900-era dates.
    """
    if col_idx <= 5:  # Jul–Dec of prior calendar year
        return date(fy - 1, 7 + col_idx, 1)
    else:  # Jan–Jun of FY calendar year
        return date(fy, col_idx - 5, 1)


def _parse_usd(value: object) -> float | None:
    """Coerce cell value to float. Returns None for blanks / None cells."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value) if value != 0 else 0.0
    text = str(value).replace("$", "").replace(",", "").strip()
    if text in ("", "None", "N/A"):
        return None
    try:
        return float(text)
    except ValueError:
        return None


Row = dict  # county, period, collections_usd, county_fips, source_url, inserted_at


def parse_fy_excel(content: bytes, fy: int, counties: list[str]) -> list[Row]:
    """Parse Form 3 Excel bytes → list of Row dicts for the requested counties."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    if SHEET_NAME not in wb.sheetnames:
        available = ", ".join(wb.sheetnames)
        raise ValueError(
            f"FY{fy}: sheet '{SHEET_NAME}' not found. Available: {available}"
        )

    ws = wb[SHEET_NAME]
    now_iso = datetime.now(timezone.utc).isoformat()
    source_url = SOURCE_URL_TMPL.format(fy=fy)

    # Derive the 12 period dates from FY + column position — do NOT parse row 9
    # cell values. Old files (FY1999, FY2001-FY2003) store month headers as small
    # Excel serials that openpyxl converts to 1900-era dates, corrupting the output.
    month_dates: list[date] = [_fy_month_date(fy, i) for i in range(12)]

    rows: list[Row] = []
    for county in counties:
        county_row_idx = _find_county_row(ws, county)
        if county_row_idx is None:
            print(
                f"  WARNING: county '{county}' not found in FY{fy} Form 3 "
                f"(col A pattern '*{county}' absent). "
                f"{'Lee County self-administers — FL DOR may have no Lee data.' if county == 'Lee' else ''}",
                file=sys.stderr,
            )
            continue

        county_row = ws[county_row_idx]
        non_zero = 0

        for month_idx, period_date in enumerate(month_dates):
            col_idx = month_idx + 1  # B=index 1, ..., M=index 12 in row tuple
            raw_val = county_row[col_idx].value
            usd = _parse_usd(raw_val)
            if usd is None:
                # Not yet published for this month — skip.
                continue
            rows.append(
                {
                    "county": county,
                    "county_fips": COUNTY_FIPS.get(county),
                    "period": period_date.isoformat(),
                    "collections_usd": usd,
                    "source_url": source_url,
                    "inserted_at": now_iso,
                }
            )
            if usd > 0:
                non_zero += 1

        if non_zero == 0 and county == "Lee":
            print(
                f"  WARNING: FY{fy} Lee County rows are all zero. "
                "FL DOR may not carry Lee data — verify leeclerk.org is the actual source.",
                file=sys.stderr,
            )

    return rows


# ── Download ──────────────────────────────────────────────────────────────────


def download_form3(fy: int, timeout: int = 180) -> bytes:
    url = FORM3_URL.format(fy=fy)
    resp = requests.get(url, timeout=timeout)
    if resp.status_code == 404:
        raise FileNotFoundError(f"FY{fy} Form 3 not found at {url}")
    resp.raise_for_status()
    return resp.content


# ── DB upsert ─────────────────────────────────────────────────────────────────


UPSERT_SQL = f"""
INSERT INTO {TABLE} (county, county_fips, period, collections_usd, source_url, inserted_at)
VALUES (%(county)s, %(county_fips)s, %(period)s, %(collections_usd)s, %(source_url)s, %(inserted_at)s)
ON CONFLICT (county, period) DO UPDATE SET
  collections_usd = EXCLUDED.collections_usd,
  source_url      = EXCLUDED.source_url,
  inserted_at     = EXCLUDED.inserted_at
"""


def upsert_rows(rows: list[Row], conn_str: str) -> int:
    """Upsert rows into fl_dor_tdt_collections. Returns count written."""
    if not rows:
        return 0
    with psycopg.connect(conn_str) as conn:
        with conn.cursor() as cur:
            cur.executemany(UPSERT_SQL, rows)
        conn.commit()
    return len(rows)


# ── Orchestration ─────────────────────────────────────────────────────────────


def run(
    fys: list[int],
    counties: list[str],
    dry_run: bool,
    conn_str: str | None,
) -> None:
    total_rows = 0
    for fy in fys:
        print(f"FY{fy}: downloading Form 3...")
        try:
            content = download_form3(fy)
        except FileNotFoundError as exc:
            print(f"  SKIP: {exc}", file=sys.stderr)
            continue
        except requests.exceptions.RequestException as exc:
            print(f"  SKIP FY{fy}: network error — {exc}", file=sys.stderr)
            continue

        rows = parse_fy_excel(content, fy, counties)
        print(f"  Parsed {len(rows)} rows for {counties}.")

        if dry_run:
            for r in rows[:5]:
                print(f"    {r}")
            if len(rows) > 5:
                print(f"    ... and {len(rows) - 5} more")
        else:
            if not conn_str:
                raise RuntimeError(
                    "DESTINATION__POSTGRES__CREDENTIALS not set — cannot write to DB."
                )
            written = upsert_rows(rows, conn_str)
            print(f"  Upserted {written} rows.")
            total_rows += written

    if not dry_run:
        print(f"Done. Total rows upserted: {total_rows}.")


# ── CLI ───────────────────────────────────────────────────────────────────────


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="FL DOR Tourist Development Tax ingest pipeline."
    )

    mode = parser.add_mutually_exclusive_group()
    mode.add_argument(
        "--backfill",
        action="store_true",
        help="Ingest all available FYs (FY1999 through current).",
    )
    mode.add_argument(
        "--current",
        action="store_true",
        help="Ingest current FY + prior FY (default for cron).",
    )
    mode.add_argument(
        "--fy",
        type=int,
        metavar="YYYY",
        help="Ingest a specific FL fiscal year only.",
    )

    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and print rows without writing to DB.",
    )
    parser.add_argument(
        "--counties",
        default=",".join(DEFAULT_COUNTIES),
        help=f"Comma-separated county names to ingest (default: {','.join(DEFAULT_COUNTIES)}).",
    )

    args = parser.parse_args(argv)
    counties = [c.strip() for c in args.counties.split(",") if c.strip()]
    conn_str = os.environ.get("DESTINATION__POSTGRES__CREDENTIALS")
    now_fy = current_fl_fy()

    if args.backfill:
        fys = fy_range(1999, now_fy)
    elif args.fy:
        fys = [args.fy]
    else:
        # --current is the default
        fys = fy_range(now_fy - 1, now_fy)

    print(
        f"fl_dor_tdt: {'dry-run ' if args.dry_run else ''}FYs {fys[0]}–{fys[-1]}, "
        f"counties={counties}"
    )
    run(fys, counties, dry_run=args.dry_run, conn_str=conn_str)
    return 0


if __name__ == "__main__":
    sys.exit(main())
